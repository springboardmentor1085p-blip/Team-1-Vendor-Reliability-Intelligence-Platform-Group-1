import csv
import io
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from ..database import get_db
from ..models import Vendor, VendorPerformance, ProcurementRequest, PurchaseOrder, Contract, User
from ..utils.dependencies import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports & Exports"])

def generate_excel_response(filename: str, headers: list, rows: list) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Data"
    
    # Write Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        
    # Write Rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = val
            
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    response = StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
    return response

def generate_pdf_response(filename: str, title: str, headers: list, rows: list) -> StreamingResponse:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        spaceAfter=12,
        textColor=colors.HexColor("#0f172a")
    )
    
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 10))
    
    cell_style = ParagraphStyle(
        name="CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10
    )
    header_style = ParagraphStyle(
        name="HeaderStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white
    )
    
    table_data = []
    table_data.append([Paragraph(str(h), header_style) for h in headers])
    for r in rows:
        table_data.append([Paragraph(str(cell or ""), cell_style) for cell in r])
        
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e293b")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")])
    ]))
    
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    
    response = StreamingResponse(buffer, media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}.pdf"
    return response

@router.get("/summary/vendors")
def get_vendor_report_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    vendors = db.query(Vendor).all()
    total = len(vendors)
    active = sum(1 for v in vendors if v.status == "Active")
    avg_score = sum(v.reliability_score for v in vendors) / total if total > 0 else 100.0
    
    risk_low = sum(1 for v in vendors if v.risk_level == "LOW")
    risk_med = sum(1 for v in vendors if v.risk_level == "MEDIUM")
    risk_high = sum(1 for v in vendors if v.risk_level == "HIGH")
    risk_crit = sum(1 for v in vendors if v.risk_level == "CRITICAL")
    
    return {
        "total_vendors": total,
        "active_vendors": active,
        "average_reliability": round(avg_score, 2),
        "risk_breakdown": {
            "LOW": risk_low,
            "MEDIUM": risk_med,
            "HIGH": risk_high,
            "CRITICAL": risk_crit
        }
    }

@router.get("/export/vendors")
def export_vendors(format: str = "csv", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    vendors = db.query(Vendor).all()
    headers = ["ID", "Name", "Category", "Status", "Contact Person", "Email", "Phone", "Reliability Score", "Risk Level", "Created At"]
    rows = []
    for v in vendors:
        rows.append([
            v.id, v.name, v.category, v.status, 
            v.contact_person or "", v.email or "", v.phone or "", 
            v.reliability_score, v.risk_level, v.created_at.strftime("%Y-%m-%d")
        ])
        
    if format == "excel":
        return generate_excel_response("vendors_report", headers, rows)
    elif format == "pdf":
        return generate_pdf_response("vendors_report", "Supplier Performance & Risk Report", headers, rows)
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=vendors_report.csv"
        return response

@router.get("/summary/procurement")
def get_procurement_report_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    reqs = db.query(ProcurementRequest).all()
    total = len(reqs)
    total_value = sum(r.estimated_cost for r in reqs)
    status_counts = {}
    for r in reqs:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1
        
    return {
        "total_requests": total,
        "total_estimated_value": total_value,
        "status_distribution": status_counts
    }

@router.get("/export/procurement")
def export_procurement(format: str = "csv", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    reqs = db.query(ProcurementRequest).all()
    headers = ["ID", "Title", "Priority", "Status", "Estimated Cost", "Requested By ID", "Approved By ID", "Vendor ID", "Created At"]
    rows = []
    for r in reqs:
        rows.append([
            r.id, r.title, r.priority, r.status, r.estimated_cost,
            r.requested_by_id, r.approved_by_id or "", r.vendor_id or "",
            r.created_at.strftime("%Y-%m-%d")
        ])
        
    if format == "excel":
        return generate_excel_response("procurement_report", headers, rows)
    elif format == "pdf":
        return generate_pdf_response("procurement_report", "Procurement Requisitions Report", headers, rows)
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=procurement_report.csv"
        return response

@router.get("/summary/purchase-orders")
def get_po_report_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    pos = db.query(PurchaseOrder).all()
    total = len(pos)
    total_val = sum(po.amount for po in pos)
    status_counts = {}
    for po in pos:
        status_counts[po.status] = status_counts.get(po.status, 0) + 1
        
    return {
        "total_pos": total,
        "total_po_value": total_val,
        "status_distribution": status_counts
    }

@router.get("/export/purchase-orders")
def export_purchase_orders(format: str = "csv", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    pos = db.query(PurchaseOrder).all()
    headers = ["ID", "PO Number", "Procurement Request ID", "Vendor ID", "Amount", "Status", "Order Date", "Expected Delivery Date", "Actual Delivery Date", "Invoice Status"]
    rows = []
    for po in pos:
        rows.append([
            po.id, po.po_number, po.procurement_request_id, po.vendor_id, po.amount, po.status,
            po.order_date.strftime("%Y-%m-%d"), po.expected_delivery_date.strftime("%Y-%m-%d"),
            po.actual_delivery_date.strftime("%Y-%m-%d") if po.actual_delivery_date else "",
            po.invoice_status
        ])
        
    if format == "excel":
        return generate_excel_response("purchase_orders_report", headers, rows)
    elif format == "pdf":
        return generate_pdf_response("purchase_orders_report", "Purchase Orders Workflow Report", headers, rows)
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=purchase_orders_report.csv"
        return response

@router.get("/summary/contracts")
def get_contracts_report_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    contracts = db.query(Contract).all()
    total = len(contracts)
    total_val = sum(c.value for c in contracts)
    status_counts = {}
    compliance_counts = {}
    for c in contracts:
        status_counts[c.status] = status_counts.get(c.status, 0) + 1
        compliance_counts[c.compliance_status] = compliance_counts.get(c.compliance_status, 0) + 1
        
    return {
        "total_contracts": total,
        "total_contract_value": total_val,
        "status_distribution": status_counts,
        "compliance_distribution": compliance_counts
    }

@router.get("/export/contracts")
def export_contracts(format: str = "csv", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    contracts = db.query(Contract).all()
    headers = ["ID", "Contract Number", "Vendor ID", "Title", "Value", "Start Date", "Expiry Date", "Status", "Compliance Status", "Certification Details"]
    rows = []
    for c in contracts:
        rows.append([
            c.id, c.contract_number, c.vendor_id, c.title, c.value,
            c.start_date.strftime("%Y-%m-%d"), c.expiry_date.strftime("%Y-%m-%d"),
            c.status, c.compliance_status, c.certification_details or ""
        ])
        
    if format == "excel":
        return generate_excel_response("contracts_report", headers, rows)
    elif format == "pdf":
        return generate_pdf_response("contracts_report", "Contracts Compliance Report", headers, rows)
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=contracts_report.csv"
        return response
